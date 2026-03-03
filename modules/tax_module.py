#所得税自动化填底稿,仅填写损益类科目

import pandas as pd
import openpyxl
from openpyxl.styles import Font
import os
from datetime import datetime
from PyQt6.QtWidgets import *
from PyQt6.QtCore import *


class TaxWorker(QThread):
    log_sig = pyqtSignal(str)
    prog_sig = pyqtSignal(int)
    finish_sig = pyqtSignal(bool, str, str)

    def __init__(self, src, tpl, sn):
        super().__init__()
        self.src, self.tpl, self.sn = src, tpl, sn

    def run(self):
        output_path = ""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = os.path.join(os.path.dirname(self.src), f"1所得税鉴证表_结果_{timestamp}.xlsm")

            self.log_sig.emit("🚀 正在读取并分析序时账...")
            self.prog_sig.emit(5)

            # 3. 增加模糊判定规则
            COL_ALIASES = {
                '一级科目': ['一级', '1级科目', '1级'],
                '二级科目': ['二级', '2级科目', '2级'],
                '三级科目': ['三级', '3级科目', '3级'],
                '四级科目': ['四级', '4级科目', '4级'],
                '借方': ['借', '借方金额'],
                '贷方': ['贷', '贷方金额']
            }

            # 自动寻找表头逻辑
            found_header = None
            for h_index in [0, 1, 2]:
                test_df = pd.read_excel(self.src, sheet_name=self.sn, header=h_index, nrows=5)
                if any(col in test_df.columns for col in ['一级科目'] + COL_ALIASES['一级科目']):
                    found_header = h_index
                    break

            if found_header is None:
                self.log_sig.emit("❌ 错误：未找到核心列名。")
                self.log_sig.emit("📝 注意：正确的列名需要按照以下格式填写：\n"
                                  "[\"机构\", \"年\", \"月\", \"凭证号\", \"摘要\", \"一级科目\", \"二级科目\", \"三级科目\", \"四级科目\", \"借方\", \"贷方\"]")
                self.finish_sig.emit(False, "表头识别失败，请检查序时账列名。", "")
                return

            df_all = pd.read_excel(self.src, sheet_name=self.sn, header=found_header)

            actual_cols = df_all.columns.tolist()
            rename_dict = {}
            for standard_name, aliases in COL_ALIASES.items():
                for alias in aliases:
                    if alias in actual_cols:
                        rename_dict[alias] = standard_name
                        break
            if rename_dict:
                df_all = df_all.rename(columns=rename_dict)

            if "机构" not in df_all.columns:
                self.log_sig.emit("⚠️ 原始数据中未发现'机构'列，将自动留空。")
                df_all["机构"] = ""

            source_cols = ["机构", "年", "月", "凭证号", "摘要", "一级科目", "二级科目", "三级科目", "四级科目", "借方",
                           "贷方"]
            for col in source_cols:
                if col not in df_all.columns:
                    df_all[col] = ""

            self.prog_sig.emit(20)

            df_all["一级科目"] = df_all["一级科目"].astype(str).str.strip()
            df_all["借方"] = pd.to_numeric(df_all["借方"], errors='coerce').fillna(0)
            df_all["贷方"] = pd.to_numeric(df_all["贷方"], errors='coerce').fillna(0)

            sheet_mapping = {
                "营业外收入": "2.1.4.1", "资产处置损益": "2.1.5.1", "其他收益": "2.1.6.1",
                "营业外支出": "2.2.4.1", "销售费用": "2.4.1", "管理费用": "2.5.1",
                "财务费用": "2.6.1", "资产减值损失": "2.7.1", "公允价值变动损益": "2.8.1",
                "投资收益": "2.9.1", "研发费用": "2.10.1", "研发支出": "2.10.1",
                "生产成本": "2.11.1", "制造费用": "2.11.1", "在建工程": "2.11.1",
                "工程施工": "2.11.1", "合同履约成本": "2.11.1", "信用减值损失": "2.12.1",
            }
            new_sheet_config = {
                "主营业务收入": "2.1.1", "主营业务成本": "2.2.1", "营业收入": "2.1.1", "营业成本": "2.2.1",
                "其他业务收入": "2.1.2", "其他业务成本": "2.2.2", "其他业务支出": "2.2.2"
            }
            DEBIT_ONLY = ["营业外支出", "销售费用", "管理费用", "财务费用", "资产减值损失", "研发费用", "研发支出",
                          "生产成本", "制造费用", "在建工程", "信用减值损失", "主营业务成本", "营业成本",
                          "其他业务成本", "其他业务支出"]
            CREDIT_ONLY = ["营业外收入", "资产处置损益", "其他收益", "公允价值变动损益", "投资收益", "主营业务收入",
                           "营业收入", "其他业务收入"]

            wb = openpyxl.load_workbook(self.tpl, keep_vba=True)
            STANDARD_FONT = Font(name="宋体", size=9)

            cleaned_sheets = set()
            subjects_to_process = list(sheet_mapping.keys()) + list(new_sheet_config.keys())
            total_subjects = len(subjects_to_process)

            for i, subject in enumerate(subjects_to_process):
                mask_subject = df_all["一级科目"].str.contains(subject, na=False)
                mask_filter = (df_all["借方"] != 0) if subject in DEBIT_ONLY else (
                    (df_all["贷方"] != 0) if subject in CREDIT_ONLY else True)

                df_sub = df_all[mask_subject & mask_filter].copy()
                self.prog_sig.emit(20 + int((i / total_subjects) * 70))

                if df_sub.empty: continue

                target_name = sheet_mapping.get(subject, subject)
                if "其他业务" in subject: target_name = "其他业务成本" if any(
                    x in subject for x in ["成本", "支出"]) else "其他业务收入"
                if "营业收入" in subject: target_name = "主营业务收入"
                if "营业成本" in subject: target_name = "主营业务成本"

                if target_name not in wb.sheetnames:
                    if subject in new_sheet_config:
                        ws = wb.create_sheet(target_name)
                    else:
                        continue
                else:
                    ws = wb[target_name]

                if target_name not in cleaned_sheets:
                    if ws.max_row >= 2: ws.delete_rows(2, ws.max_row)
                    cleaned_sheets.add(target_name)

                self.log_sig.emit(f"📝 正在提取科目: {subject} -> {target_name}")
                data_list = df_sub[source_cols].values.tolist()
                start_row = ws.max_row + 1
                for r_idx, r_data in enumerate(data_list):
                    for c_idx, val in enumerate(r_data, 3):
                        cell = ws.cell(start_row + r_idx, c_idx, val)
                        cell.font = STANDARD_FONT

            self.log_sig.emit("💾 正在保存生成的结果文件...")
            wb.save(output_path)
            self.prog_sig.emit(100)
            self.finish_sig.emit(True, "底稿生成成功！", output_path)
            self.log_sig.emit("💾 结果文件已保存")
        except Exception as e:
            self.finish_sig.emit(False, f"执行失败：{str(e)}", "")


class IncomeTaxPage(QWidget):
    """插件界面类，保持原有功能逻辑不动"""
    def __init__(self):
        super().__init__()
        self.last_output_dir = ""
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        group = QGroupBox("所得税鉴证底稿自动提取")
        grid = QGridLayout(group)

        grid.addWidget(QLabel("源序时账:"), 0, 0)
        self.src_edit = QLineEdit()
        grid.addWidget(self.src_edit, 0, 1)
        btn_src = QPushButton("选择文件")
        btn_src.clicked.connect(lambda: self.select_path(self.src_edit))
        grid.addWidget(btn_src, 0, 2)

        grid.addWidget(QLabel("底稿模板:"), 1, 0)
        self.tpl_edit = QLineEdit()
        grid.addWidget(self.tpl_edit, 1, 1)
        btn_tpl = QPushButton("选择文件")
        btn_tpl.clicked.connect(lambda: self.select_path(self.tpl_edit))
        grid.addWidget(btn_tpl, 1, 2)

        grid.addWidget(QLabel("序时账表名:"), 2, 0)
        self.sn_edit = QLineEdit("序时账")
        grid.addWidget(self.sn_edit, 2, 1)

        grid.addWidget(QLabel("处理进度:"), 3, 0)
        self.pbar = QProgressBar()
        self.pbar.setTextVisible(True)
        grid.addWidget(self.pbar, 3, 1)

        layout.addWidget(group)

        btn_layout = QHBoxLayout()
        self.run_btn = QPushButton("🚀 开始执行任务")
        self.run_btn.setFixedHeight(45)
        self.run_btn.clicked.connect(self.start_task)
        self.open_dir_btn = QPushButton("📂 打开结果目录")
        self.open_dir_btn.setFixedHeight(45)
        self.open_dir_btn.setEnabled(False)
        self.open_dir_btn.clicked.connect(self.open_result_folder)
        btn_layout.addWidget(self.run_btn)
        btn_layout.addWidget(self.open_dir_btn)
        layout.addLayout(btn_layout)

        self.log = QTextEdit()
        self.log.setReadOnly(True)
        layout.addWidget(self.log)

    def select_path(self, edit):
        path, _ = QFileDialog.getOpenFileName(self, "选择文件", "", "Excel (*.xlsx *.xlsm)")
        if path: edit.setText(path)

    def start_task(self):
        if not self.src_edit.text() or not self.tpl_edit.text():
            QMessageBox.warning(self, "提示", "请完整选择文件！")
            return
        self.run_btn.setEnabled(False)
        self.pbar.setValue(0)
        self.log.clear()
        self.worker = TaxWorker(self.src_edit.text(), self.tpl_edit.text(), self.sn_edit.text())
        self.worker.log_sig.connect(self.log.append)
        self.worker.prog_sig.connect(self.pbar.setValue)
        self.worker.finish_sig.connect(self.on_finish)
        self.worker.start()

    def on_finish(self, success, msg, output_path):
        self.run_btn.setEnabled(True)
        if success:
            self.last_output_dir = os.path.dirname(output_path)
            self.open_dir_btn.setEnabled(True)
            QMessageBox.information(self, "完成", msg)
        else:
            QMessageBox.critical(self, "失败", msg)

    def open_result_folder(self):
        if self.last_output_dir: os.startfile(self.last_output_dir)


# ==================== 插件接口 ====================
class Plugin:
    plugin_name = "所得税底稿自动化"
    plugin_version = "1.0.0"

    def get_widget(self):
        return IncomeTaxPage()


def register_plugin():
    return Plugin()


# ==================== 独立运行测试程序 ====================
def _test_run():
    """
    当此脚本作为主程序运行时，将启动独立测试模式
    """
    import sys

    # 1. 设置高分屏适配，防止界面模糊
    if hasattr(Qt, 'HighDpiScaleFactorRoundingPolicy'):
        QApplication.setHighDpiScaleFactorRoundingPolicy(Qt.HighDpiScaleFactorRoundingPolicy.PassThrough)

    app = QApplication(sys.argv)

    # 3. 实例化并配置测试窗口
    test_window = IncomeTaxPage()
    test_window.setWindowTitle("独立测试模式 - 所得税底稿自动化")
    test_window.resize(1000, 750)  # 设置一个适合底稿提取操作的舒适大小

    # 将窗口居中显示
    frame_gm = test_window.frameGeometry()
    screen_center = app.primaryScreen().availableGeometry().center()
    frame_gm.moveCenter(screen_center)
    test_window.move(frame_gm.topLeft())

    test_window.show()

    print("🚀 所得税底稿自动化模块已进入独立测试运行模式")
    sys.exit(app.exec())


if __name__ == "__main__":
    _test_run()