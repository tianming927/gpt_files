#研发工时分配工具,根据工资表,项目人员及研发工资明细账对工时进行分配

import sys
import os
import time
import shutil
import threading
import pandas as pd
import numpy as np
import pulp
from openpyxl import load_workbook
from collections import defaultdict
from PyQt6.QtWidgets import *
from PyQt6.QtCore import *
from PyQt6.QtGui import *


# ====================== 核心计算逻辑类 ======================
class RDProcessor(QObject):
    log_signal = pyqtSignal(str)
    finished_signal = pyqtSignal(bool, str)

    def __init__(self, file_path, template_salary, template_hours, time_limit):
        super().__init__()
        self.file_path = file_path
        self.template_salary = template_salary
        self.template_hours = template_hours
        self.time_limit = time_limit

    def run(self):
        try:
            base_dir = os.path.dirname(self.file_path)
            salary_output_file = os.path.join(base_dir, "1.工资分配结果表.xlsx")
            hours_output_file = os.path.join(base_dir, "2.工时分配结果表.xlsx")

            # 1. 数据读取
            self.log_signal.emit("正在读取Excel数据源...")
            sheet2 = pd.read_excel(self.file_path, sheet_name="3.基本工资表", header=None)

            employee_names = {}
            for i in range(2, len(sheet2)):
                k = sheet2.iloc[i, 0]
                n = sheet2.iloc[i, 1]
                if pd.notna(k) and pd.notna(n):
                    clean_id = str(k).split('.')[0].strip()
                    employee_names[clean_id] = str(n).strip()

            sheet1 = pd.read_excel(self.file_path, sheet_name="2.各项目每月研发金额汇总及核验", header=None)
            project_budgets = {}
            for i in range(2, len(sheet1)):
                k = sheet1.iloc[i, 0]
                if pd.notna(k):
                    project_budgets[str(k)] = [
                        round(float(v), 2) if pd.notna(v) and isinstance(v, (int, float)) else 0.0 for v in
                        sheet1.iloc[i, 2:14]]

            employee_salaries = {}
            salary_hourly = {}
            for i in range(2, len(sheet2)):
                k = sheet2.iloc[i, 0]
                if pd.notna(k):
                    clean_k = str(k).split('.')[0].strip()
                    employee_salaries[clean_k] = [
                        round(float(v), 2) if pd.notna(v) and isinstance(v, (int, float)) else 0.0 for v in
                        sheet2.iloc[i, 2:14]]
                    salary_hourly[clean_k] = [
                        round(float(v), 2) if pd.notna(v) and isinstance(v, (int, float)) else 0.0
                        for v in sheet2.iloc[i, 29:41]]

            sheet3 = pd.read_excel(self.file_path, sheet_name="取数表_勿删", header=None)
            project_assignments = {}
            p_keys = list(project_budgets.keys())
            for i, k in enumerate(p_keys):
                row = i + 1
                if row < len(sheet3):
                    assigned = []
                    col = 2
                    while col < len(sheet3.columns) and pd.notna(sheet3.iloc[row, col]):
                        emp_id_clean = str(sheet3.iloc[row, col]).strip().split('.')[0]
                        assigned.append(emp_id_clean)
                        col += 1
                    project_assignments[k] = assigned

            all_emps = list(employee_salaries.keys())
            monthly_h_list = [23, 18, 21, 22, 21, 19, 23, 22, 21, 19, 21, 22]
            start_total = time.time()

            for month in range(1, 13):
                m_idx = month - 1
                self.log_signal.emit(f">>> 处理第 {month} 月数据...")
                res_df, s_vars = self.optimize_month(month, monthly_h_list[m_idx], all_emps, p_keys, project_budgets,
                                                     employee_salaries, salary_hourly, project_assignments)
                if res_df is not None:
                    hours_df = pd.DataFrame(index=all_emps, columns=p_keys).fillna(0)
                    for p in p_keys:
                        for e in project_assignments.get(p, []):
                            val = s_vars.get((p, e))
                            if val and val.varValue:
                                hours_df.at[e, p] = int(val.varValue)

                    salary_df = res_df.T.copy()
                    salary_df.index = salary_df.index.map(lambda x: str(x).split('.')[0].strip())
                    self.write_excel(month, res_df, self.template_salary, salary_output_file, employee_names, p_keys, False)
                    self.write_excel(month, hours_df, self.template_hours, hours_output_file, employee_names, p_keys, True)
                elapsed = time.time() - start_total
                avg = elapsed / month
                remain = avg * (12 - month)
                self.log_signal.emit(f"    {month}月完成。预计还需: {remain:.1f} 秒")

            self.finished_signal.emit(True, f"结果保存在: {base_dir}")
        except Exception as e:
            self.finished_signal.emit(False, str(e))

    def optimize_month(self, month, m_days, all_emps, p_keys, p_budgets, e_sals, s_hourly, p_assigns):
        m_idx = month - 1
        total_h = m_days * 8
        model = pulp.LpProblem(f"M_{month}", pulp.LpMinimize)
        s_vars = {}
        d_p = pulp.LpVariable.dicts("dp", p_keys, 0)
        d_m = pulp.LpVariable.dicts("dm", p_keys, 0)

        for p in p_keys:
            budget = p_budgets[p][m_idx]
            if budget <= 0: continue
            for e in p_assigns.get(p, []):
                if e in all_emps:
                    emp_sal = e_sals[e][m_idx]
                    emp_hr = s_hourly[e][m_idx]
                    max_h = min(int(emp_sal / emp_hr), total_h) if emp_hr > 0 else 0
                    s_vars[(p, e)] = pulp.LpVariable(f"v_{month}_{p}_{e}", 0, max_h, cat='Integer')

        for e in all_emps:
            rel = [s_vars[(p, e)] * s_hourly[e][m_idx] for p in p_keys if (p, e) in s_vars]
            if rel: model += pulp.lpSum(rel) <= e_sals[e][m_idx]

        for p in p_keys:
            budget = p_budgets[p][m_idx]
            if budget > 0:
                rel = [s_vars[(p, e)] * s_hourly[e][m_idx] for e in all_emps if (p, e) in s_vars]
                if rel: model += pulp.lpSum(rel) + d_m[p] - d_p[p] == budget

        model += pulp.lpSum([d_p[p] + d_m[p] for p in p_keys])
        solver = pulp.PULP_CBC_CMD(msg=False, timeLimit=self.time_limit)
        model.solve(solver)

        res = {p: {e: 0 for e in all_emps} for p in p_keys}
        for (p, e), var in s_vars.items():
            if var.varValue:
                res[p][e] = round(var.varValue * s_hourly[e][m_idx], 2)

        return pd.DataFrame(res), s_vars

    def write_excel(self, month, df, tpl, out, names, p_keys, is_h):
        sheet_n = f"{month}月"
        if month == 1 or not os.path.exists(out): shutil.copy2(tpl, out)
        wb = load_workbook(out)
        if sheet_n not in wb.sheetnames: wb.create_sheet(sheet_n)
        ws = wb[sheet_n]

        active_emps = [e for e in df.index if df.loc[e].sum() > 0]
        max_r = ws.max_row if ws.max_row >= 6 else 6
        for r in range(6, max_r + 1):
            ws.cell(r, 2).value = None
            ws.cell(r, 3).value = None
            for c in range(7, ws.max_column + 1):
                ws.cell(r, c).value = None

        col_map = {}
        for c in range(7, ws.max_column + 1):
            p_id = str(ws.cell(2, c).value).strip()
            if p_id in p_keys: col_map[p_id] = c

        for i, eid in enumerate(active_emps):
            current_row = 6 + i
            clean_eid = str(eid).split('.')[0].strip()
            lookup_id = clean_eid if clean_eid in df.index else eid
            ws.cell(current_row, 2, names.get(clean_eid, "未知"))
            ws.cell(current_row, 3, clean_eid)
            for p, c_idx in col_map.items():
                if p in df.columns and lookup_id in df.index:
                    val = df.at[lookup_id, p]
                    if pd.notna(val) and val > 0:
                        ws.cell(current_row, c_idx, int(val) if is_h else val)
        wb.save(out)


# ====================== PyQt6 UI 组件类 ======================
class RDPage(QWidget):
    """插件界面类，保持原有功能逻辑不动，只修改类名以适配主程序"""
    def __init__(self):
        super().__init__()
        self.output_dir = ""
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        header = QLabel("研发工资/工时自动化分配模块")
        header.setStyleSheet("font-size: 18px; font-weight: bold; color: #2c3e50;")
        layout.addWidget(header)

        file_box = QGroupBox("1. 文件设置")
        file_layout = QVBoxLayout(file_box)
        self.path_edit = self._create_file_selector(file_layout, "数据源文件:   ", "请选择测试文件.xlsx",
                                                    self.select_data_file)
        self.tpl_salary_edit = self._create_file_selector(file_layout, "工资分配模板:", "请选择工资分配空白模板.xlsx",
                                                          self.select_tpl_salary)
        self.tpl_hours_edit = self._create_file_selector(file_layout, "工时分配模板:", "请选择工时分配空白模板.xlsx",
                                                         self.select_tpl_hours)
        layout.addWidget(file_box)

        param_box = QGroupBox("2. 参数设置")
        param_layout = QHBoxLayout(param_box)
        param_layout.addWidget(QLabel("求解时间限制 (秒):"))
        self.time_spin = QSpinBox()
        self.time_spin.setRange(1, 3600)
        self.time_spin.setValue(10)
        param_layout.addWidget(self.time_spin)
        param_layout.addStretch()
        layout.addWidget(param_box)

        log_box = QGroupBox("3. 处理进度")
        log_layout = QVBoxLayout(log_box)
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setStyleSheet("background-color: #f8f9fa; font-family: 'Consolas';")
        log_layout.addWidget(self.log_text)
        layout.addWidget(log_box)

        btn_layout = QHBoxLayout()
        self.btn_start = QPushButton("🚀 开始执行任务")
        self.btn_start.setFixedHeight(40)
        self.btn_start.setStyleSheet("background-color: #3498db; color: white; font-weight: bold;")
        self.btn_start.clicked.connect(self.start_task)
        self.btn_folder = QPushButton("📂 打开结果文件夹")
        self.btn_folder.setFixedHeight(40)
        self.btn_folder.setEnabled(False)
        self.btn_folder.clicked.connect(self.open_folder)
        btn_layout.addWidget(self.btn_start)
        btn_layout.addWidget(self.btn_folder)
        layout.addLayout(btn_layout)

    def _create_file_selector(self, parent_layout, label_text, placeholder, slot):
        h_layout = QHBoxLayout()
        h_layout.addWidget(QLabel(label_text))
        line_edit = QLineEdit()
        line_edit.setPlaceholderText(placeholder)
        btn = QPushButton("浏览")
        btn.clicked.connect(slot)
        h_layout.addWidget(line_edit)
        h_layout.addWidget(btn)
        parent_layout.addLayout(h_layout)
        return line_edit

    def select_data_file(self):
        file, _ = QFileDialog.getOpenFileName(self, "选择数据源", "", "Excel Files (*.xlsx *.xls)")
        if file:
            self.path_edit.setText(file)
            self.output_dir = os.path.dirname(file)

    def select_tpl_salary(self):
        start = self.output_dir if self.output_dir else ""
        file, _ = QFileDialog.getOpenFileName(self, "选择工资模板", start, "Excel Files (*.xlsx *.xls)")
        if file: self.tpl_salary_edit.setText(file)

    def select_tpl_hours(self):
        start = self.output_dir if self.output_dir else ""
        file, _ = QFileDialog.getOpenFileName(self, "选择工时模板", start, "Excel Files (*.xlsx *.xls)")
        if file: self.tpl_hours_edit.setText(file)

    def open_folder(self):
        if self.output_dir: os.startfile(self.output_dir)

    def start_task(self):
        path, tpl_s, tpl_h = self.path_edit.text(), self.tpl_salary_edit.text(), self.tpl_hours_edit.text()
        if not all([path, tpl_s, tpl_h]) or not all([os.path.exists(p) for p in [path, tpl_s, tpl_h]]):
            QMessageBox.warning(self, "错误", "请确保所有文件已选择且路径正确！")
            return
        self.btn_start.setEnabled(False)
        self.btn_start.setText("正在计算...")
        self.log_text.clear()
        self.thread = QThread()
        self.worker = RDProcessor(path, tpl_s, tpl_h, self.time_spin.value())
        self.worker.moveToThread(self.thread)
        self.thread.started.connect(self.worker.run)
        self.worker.log_signal.connect(self.update_log)
        self.worker.finished_signal.connect(self.task_done)
        self.thread.start()

    def update_log(self, text):
        self.log_text.append(text)

    def task_done(self, success, msg):
        self.thread.quit()
        self.btn_start.setEnabled(True)
        self.btn_start.setText("开始执行任务")
        if success:
            self.btn_folder.setEnabled(True)
            QMessageBox.information(self, "成功", "计算完成！")
        else:
            QMessageBox.critical(self, "失败", f"发生错误:\n{msg}")


# ==================== 插件接口 ====================
class Plugin:
    plugin_name = "研发工资&工时分配"
    plugin_version = "1.0.0"

    def get_widget(self):
        return RDPage()


def register_plugin():
    return Plugin()


# ==================== 独立运行测试程序 ====================
def _test_run():
    """
    当此脚本作为主程序运行时，将启动独立测试模式
    """
    import sys
    # 设置高分屏适配
    if hasattr(Qt, 'HighDpiScaleFactorRoundingPolicy'):
        QApplication.setHighDpiScaleFactorRoundingPolicy(Qt.HighDpiScaleFactorRoundingPolicy.PassThrough)

    app = QApplication(sys.argv)

    # 设置全局样式（可选，模拟主程序的视觉风格）
    app.setStyleSheet("""
        QWidget { font-family: 'Microsoft YaHei'; }
        QGroupBox { font-weight: bold; }
    """)

    # 实例化界面类
    test_window = RDPage()
    test_window.setWindowTitle("独立测试模式 - 研发工资&工时分配")
    test_window.resize(900, 700)  # 设置一个默认的舒适大小
    test_window.show()

    print("🧪 模块已进入独立测试运行模式")
    sys.exit(app.exec())


if __name__ == "__main__":
    _test_run()
